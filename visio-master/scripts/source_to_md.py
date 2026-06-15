#!/usr/bin/env python3
"""
source_to_md.py — Unified source ingestion for the visio-master skill.

Converts the source formats that feed Visio diagrams (data tables, org
rosters, BOM lists, prose briefings) into Markdown that Architect reads
during the Eight Confirmations and Drafter quotes into ShapeSheet text.

Supported inputs
    .csv / .tsv   stdlib csv (delimiter sniffed)
    .txt          stdlib text (multi-encoding fallback)
    .md           Markdown passthrough with line-ending normalisation
    .xlsx / .xlsm openpyxl (optional)
    .pdf          PyMuPDF / fitz (optional)

Optional dependencies (declared, NOT installed by this script)
    openpyxl >= 3.1     for .xlsx / .xlsm
    PyMuPDF  >= 1.24    for .pdf
    pywin32             only used to recognise pywintypes.com_error in
                        the top-level CLI guard; absence is fine.
    vsdx                reserved for downstream Visio scripts; lazily
                        imported here so this module loads without it.

Output convention
    <project_path>/sources/<input.stem>.md   when --project is given
    <input>.md                               otherwise

Usage
    python source_to_md.py convert org_roster.xlsx --project ./drawing
    python source_to_md.py convert spec.pdf -o sources/spec.md
    python source_to_md.py batch ./inputs --project ./drawing
    python source_to_md.py inspect data.csv

Exit codes
    0   success
    1   recoverable failure (missing dep / unknown format / file missing)
    2   I/O or COM-style error during conversion
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

# Optional third-party imports — degrade gracefully ----------------------

try:
    from openpyxl import load_workbook  # type: ignore[import-not-found]
    from openpyxl.utils import get_column_letter  # type: ignore[import-not-found]
    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - reported at conversion time
    load_workbook = None  # type: ignore[assignment]
    get_column_letter = None  # type: ignore[assignment]
    _HAS_OPENPYXL = False

try:
    import fitz  # type: ignore[import-not-found]
    _HAS_FITZ = True
except ImportError:  # pragma: no cover - reported at conversion time
    fitz = None  # type: ignore[assignment]
    _HAS_FITZ = False

try:  # pragma: no cover - parity import only
    import pywintypes  # type: ignore[import-not-found]
    _HAS_PYWIN32 = True
except ImportError:  # pragma: no cover
    pywintypes = None  # type: ignore[assignment]
    _HAS_PYWIN32 = False

try:  # pragma: no cover - parity import only
    import vsdx  # type: ignore[import-not-found]  # noqa: F401
    _HAS_VSDX = True
except ImportError:  # pragma: no cover
    _HAS_VSDX = False


# Format registry --------------------------------------------------------

CSV_FORMATS = {".csv"}
TSV_FORMATS = {".tsv"}
TEXT_FORMATS = {".txt"}
MARKDOWN_FORMATS = {".md", ".markdown"}
EXCEL_FORMATS = {".xlsx", ".xlsm"}
LEGACY_EXCEL_FORMATS = {".xls"}
PDF_FORMATS = {".pdf"}

ALL_SUPPORTED = (
    CSV_FORMATS | TSV_FORMATS | TEXT_FORMATS
    | MARKDOWN_FORMATS | EXCEL_FORMATS | PDF_FORMATS
)

CONTROL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
TEXT_ENCODINGS = ("utf-8-sig", "utf-8", "gb18030", "cp1252", "latin-1")


# Shared helpers ---------------------------------------------------------

def _format_size(size: float) -> str:
    for unit in ("B", "KB", "MB"):
        if size < 1024:
            return f"{size:.0f} {unit}"
        size /= 1024
    return f"{size:.1f} GB"


def _is_empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _markdown_escape(value: str) -> str:
    value = value.replace("\\", "\\\\").replace("|", "\\|")
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    return re.sub(r"\s*\n\s*", "<br>", value).strip()


def _format_cell_value(value: Any) -> str:
    if _is_empty(value):
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, float):
        return _markdown_escape(f"{value:g}")
    return _markdown_escape(str(value))


def _is_numeric_value(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _read_text_with_fallback(path: Path) -> str:
    for enc in TEXT_ENCODINGS:
        try:
            return path.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("text", b"", 0, 1, f"cannot decode {path}")


def _resolve_output_path(
    input_file: Path, output: str | None, project: str | None,
) -> Path:
    if output:
        return Path(output)
    if project:
        return Path(project) / "sources" / f"{input_file.stem}.md"
    return input_file.with_suffix(".md")


def _emit_summary(
    out_file: Path, input_file: Path, fmt: str, extra: dict[str, Any] | None,
) -> dict[str, Any]:
    summary: dict[str, Any] = {
        "status": "ok",
        "input": str(input_file),
        "output": str(out_file),
        "format": fmt,
        "bytes": out_file.stat().st_size if out_file.exists() else 0,
    }
    if extra:
        summary.update(extra)
    print(json.dumps(summary, ensure_ascii=False))
    return summary


def _print_error(msg: str) -> None:
    print(f"[ERROR] {msg}", file=sys.stderr)


# Tabular helpers (CSV / TSV / XLSX) -------------------------------------

def _trim_trailing_empty(row: list[Any]) -> list[Any]:
    trimmed = list(row)
    while trimmed and _is_empty(trimmed[-1]):
        trimmed.pop()
    return trimmed


def _column_alignments(rows: list[list[Any]]) -> list[str]:
    if not rows:
        return []
    width = len(rows[0])
    data_rows = rows[1:] if len(rows) > 1 else rows
    alignments: list[str] = []
    for col in range(width):
        values = [
            r[col] for r in data_rows
            if col < len(r) and not _is_empty(r[col])
        ]
        alignments.append(
            "---:" if values and all(_is_numeric_value(v) for v in values) else "---"
        )
    return alignments


def _rows_to_markdown_table(rows: list[list[Any]]) -> str:
    if not rows:
        return "_No tabular content found._"
    formatted = [[_format_cell_value(v) for v in r] for r in rows]
    width = max((len(r) for r in formatted), default=0)
    if width == 0:
        return "_No tabular content found._"
    formatted = [r + [""] * (width - len(r)) for r in formatted]
    sep = _column_alignments(rows) or ["---"] * width
    if len(sep) < width:
        sep += ["---"] * (width - len(sep))
    out = ["| " + " | ".join(formatted[0]) + " |", "| " + " | ".join(sep) + " |"]
    for r in formatted[1:]:
        out.append("| " + " | ".join(r) + " |")
    return "\n".join(out)


def _truncate_rows(
    rows: list[list[Any]], max_rows: int, max_cols: int,
) -> tuple[list[list[Any]], bool, bool]:
    rows_trunc = max_rows > 0 and len(rows) > max_rows
    if rows_trunc:
        rows = rows[:max_rows]
    cols_trunc = False
    if max_cols > 0:
        new: list[list[Any]] = []
        for r in rows:
            if len(r) > max_cols:
                r = r[:max_cols]
                cols_trunc = True
            new.append(r)
        rows = new
    return rows, rows_trunc, cols_trunc


# CSV / TSV --------------------------------------------------------------

def _detect_dialect(sample: str, default_delim: str):
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        class _Fallback(csv.Dialect):
            delimiter = default_delim
            quotechar = '"'
            doublequote = True
            skipinitialspace = False
            lineterminator = "\n"
            quoting = csv.QUOTE_MINIMAL
        return _Fallback


def _convert_csv(
    input_file: Path, out_file: Path,
    max_rows: int, max_cols: int, is_tsv: bool,
) -> dict[str, Any]:
    text = _read_text_with_fallback(input_file)
    dialect = _detect_dialect(text[:8192], "\t" if is_tsv else ",")
    rows = [list(r) for r in csv.reader(text.splitlines(), dialect=dialect)]
    rows = [_trim_trailing_empty(r) for r in rows]
    rows = [r for r in rows if any(not _is_empty(c) for c in r)]
    rows, rows_trunc, cols_trunc = _truncate_rows(rows, max_rows, max_cols)
    width = max((len(r) for r in rows), default=0)
    rows = [r + [""] * (width - len(r)) for r in rows]

    label = "TSV" if is_tsv else "CSV"
    lines = [
        f"# {label} Source: {input_file.name}", "",
        "## Summary", "",
        f"- Rows: {len(rows)}",
        f"- Columns: {width}", "",
    ]
    if rows_trunc or cols_trunc:
        notes = []
        if rows_trunc:
            notes.append(f"rows limited to {max_rows}")
        if cols_trunc:
            notes.append(f"columns limited to {max_cols}")
        lines += [f"> Truncated: {', '.join(notes)}.", ""]
    lines.append(_rows_to_markdown_table(rows))
    out_file.parent.mkdir(parents=True, exist_ok=True)
    out_file.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    return {"rows": len(rows), "cols": width}


# Plain text / Markdown --------------------------------------------------

def _convert_text(input_file: Path, out_file: Path) -> dict[str, Any]:
    raw = _read_text_with_fallback(input_file)
    raw = raw.replace("\r\n", "\n").replace("\r", "\n")
    raw = CONTROL_CHARS_RE.sub("", raw).rstrip() + "\n"
    out_file.parent.mkdir(parents=True, exist_ok=True)
    out_file.write_text(
        f"# Text Source: {input_file.name}\n\n```\n{raw}```\n",
        encoding="utf-8",
    )
    return {"chars": len(raw)}


def _convert_markdown(input_file: Path, out_file: Path) -> dict[str, Any]:
    raw = _read_text_with_fallback(input_file)
    raw = raw.replace("\r\n", "\n").replace("\r", "\n")
    raw = CONTROL_CHARS_RE.sub("", raw).rstrip() + "\n"
    out_file.parent.mkdir(parents=True, exist_ok=True)
    out_file.write_text(raw, encoding="utf-8")
    return {"chars": len(raw)}


# Excel ------------------------------------------------------------------

def _merged_value_map(ws: Any) -> dict[tuple[int, int], Any]:
    merged: dict[tuple[int, int], Any] = {}
    for rng in ws.merged_cells.ranges:
        v = ws.cell(rng.min_row, rng.min_col).value
        if _is_empty(v):
            continue
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged[(r, c)] = v
    return merged


def _excel_bounds(
    ws: Any, merged: dict[tuple[int, int], Any],
) -> tuple[int, int, int, int] | None:
    min_r = min_c = max_r = max_c = None
    for row in ws.iter_rows():
        for cell in row:
            if _is_empty(cell.value):
                continue
            min_r = cell.row if min_r is None else min(min_r, cell.row)
            max_r = cell.row if max_r is None else max(max_r, cell.row)
            min_c = cell.column if min_c is None else min(min_c, cell.column)
            max_c = cell.column if max_c is None else max(max_c, cell.column)
    for r, c in merged:
        min_r = r if min_r is None else min(min_r, r)
        max_r = r if max_r is None else max(max_r, r)
        min_c = c if min_c is None else min(min_c, c)
        max_c = c if max_c is None else max(max_c, c)
    if min_r is None:
        return None
    return min_r, min_c, max_r, max_c


def _extract_excel_rows(
    ws: Any, bounds: tuple[int, int, int, int],
    merged: dict[tuple[int, int], Any],
    max_rows: int, max_cols: int,
) -> tuple[list[list[Any]], bool, bool]:
    min_r, min_c, max_r, max_c = bounds
    row_lim, col_lim = max_r, max_c
    rows_trunc = max_rows > 0 and (max_r - min_r + 1) > max_rows
    cols_trunc = max_cols > 0 and (max_c - min_c + 1) > max_cols
    if rows_trunc:
        row_lim = min_r + max_rows - 1
    if cols_trunc:
        col_lim = min_c + max_cols - 1

    rows: list[list[Any]] = []
    width = 0
    for r in range(min_r, row_lim + 1):
        row = []
        for c in range(min_c, col_lim + 1):
            v = ws.cell(r, c).value
            if _is_empty(v):
                v = merged.get((r, c), v)
            row.append(v)
        row = _trim_trailing_empty(row)
        width = max(width, len(row))
        rows.append(row)
    if width == 0:
        return [], rows_trunc, cols_trunc
    return [r + [""] * (width - len(r)) for r in rows], rows_trunc, cols_trunc


def _convert_excel(
    input_file: Path, out_file: Path, max_rows: int, max_cols: int,
) -> dict[str, Any]:
    if not _HAS_OPENPYXL:
        raise RuntimeError(
            "openpyxl not available. Install with: pip install 'openpyxl>=3.1'"
        )
    wb = load_workbook(input_file, data_only=True, read_only=False)
    visible = [s for s in wb.worksheets if s.sheet_state == "visible"]
    lines = [
        f"# Spreadsheet Source: {input_file.name}", "",
        "## Workbook Summary", "",
        f"- Sheets: {len(wb.worksheets)}",
        f"- Visible sheets: {', '.join(s.title for s in visible) or 'None'}", "",
        "> Note: Formula cells are exported as cached values. "
        "This converter does not recalculate formulas.", "",
    ]
    if not visible:
        lines += ["_No visible sheets found._", ""]
    sheet_count = 0
    for ws in visible:
        merged = _merged_value_map(ws)
        bounds = _excel_bounds(ws, merged)
        lines += [f"## Sheet: {ws.title}", "", f"- State: {ws.sheet_state or 'unknown'}"]
        if bounds is None:
            lines += ["", "_No content found._", ""]
            continue
        min_r, min_c, max_r, max_c = bounds
        used = (
            f"{get_column_letter(min_c)}{min_r}:"
            f"{get_column_letter(max_c)}{max_r}"
        )
        rows, rows_trunc, cols_trunc = _extract_excel_rows(
            ws, bounds, merged, max_rows, max_cols
        )
        lines += [
            f"- Used range: {used}",
            f"- Rows: {max_r - min_r + 1}",
            f"- Columns: {max_c - min_c + 1}", "",
        ]
        if rows_trunc or cols_trunc:
            notes = []
            if rows_trunc:
                notes.append(f"rows limited to {max_rows}")
            if cols_trunc:
                notes.append(f"columns limited to {max_cols}")
            lines += [f"> Truncated: {', '.join(notes)}.", ""]
        lines += [_rows_to_markdown_table(rows), ""]
        sheet_count += 1
    out_file.parent.mkdir(parents=True, exist_ok=True)
    out_file.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    return {"sheets": sheet_count}


# PDF --------------------------------------------------------------------

def _pdf_format_span(text: str, flags: int) -> str:
    text = CONTROL_CHARS_RE.sub("", text).strip()
    if not text:
        return ""
    bold, italic = bool(flags & 16), bool(flags & 2)
    if bold and italic:
        return f"***{text}***"
    if bold:
        return f"**{text}**"
    if italic:
        return f"*{text}*"
    return text


def _pdf_heading_level(size: float, body_size: float) -> int:
    if size >= body_size + 8:
        return 1
    if size >= body_size + 4:
        return 2
    if size >= body_size + 2:
        return 3
    return 0


def _pdf_body_size(doc: Any) -> float:
    from collections import Counter
    counter: Counter[float] = Counter()
    for page in doc:
        try:
            blocks = page.get_text("dict")["blocks"]
        except Exception:  # pragma: no cover - defensive
            continue
        for b in blocks:
            if b.get("type") != 0:
                continue
            for line in b.get("lines", []):
                for span in line.get("spans", []):
                    text = span.get("text", "").strip()
                    if text:
                        counter[round(span.get("size", 12.0), 1)] += len(text)
    return counter.most_common(1)[0][0] if counter else 12.0


def _pdf_render_page(page: Any, body_size: float, page_index: int) -> str:
    chunks: list[str] = []
    try:
        tabs = page.find_tables()
    except Exception:  # pragma: no cover - defensive
        tabs = None
    table_rects: list[Any] = []
    if tabs is not None:
        for tab in tabs:
            try:
                table_rects.append(fitz.Rect(tab.bbox))
                md = tab.to_markdown().strip()
                if md:
                    chunks.append(md)
            except Exception:  # pragma: no cover - defensive
                continue

    try:
        blocks = page.get_text("dict")["blocks"]
    except Exception:  # pragma: no cover - defensive
        blocks = []

    lines: list[str] = []
    prev_blank = False
    for block in blocks:
        if block.get("type") != 0:
            continue
        rect = fitz.Rect(block["bbox"])
        if any(
            (rect & tr).get_area() > 0.6 * rect.get_area() for tr in table_rects
        ):
            continue
        for line in block.get("lines", []):
            spans = line.get("spans", [])
            if not spans:
                continue
            line_size = max(
                (s.get("size", body_size) for s in spans), default=body_size
            )
            text = "".join(
                _pdf_format_span(s.get("text", ""), s.get("flags", 0))
                for s in spans
            ).strip()
            if not text:
                if not prev_blank and lines:
                    lines.append("")
                    prev_blank = True
                continue
            level = _pdf_heading_level(line_size, body_size)
            stripped = re.sub(r"\*+([^*]+)\*+", r"\1", text) if level else text
            ul = re.match(r"^(?:[•●○◦▪▸►]|[-–—]|\*)\s+(.*)$", stripped)
            ol = re.match(r"^(\d+)[.、)]\s+(.*)$", stripped)
            if level:
                lines += [f"{'#' * level} {stripped}", ""]
                prev_blank = True
            elif ul:
                lines.append(f"- {ul.group(1)}")
                prev_blank = False
            elif ol:
                lines.append(f"{ol.group(1)}. {ol.group(2)}")
                prev_blank = False
            else:
                lines.append(stripped)
                prev_blank = False

    text_block = "\n".join(lines).strip()
    if text_block:
        chunks.insert(0, text_block)
    if not chunks:
        return ""
    return f"<!-- Page {page_index} -->\n\n" + "\n\n".join(chunks)


def _convert_pdf(input_file: Path, out_file: Path) -> dict[str, Any]:
    if not _HAS_FITZ:
        raise RuntimeError(
            "PyMuPDF not available. Install with: pip install 'PyMuPDF>=1.24'"
        )
    doc = fitz.open(str(input_file))
    try:
        body_size = _pdf_body_size(doc)
        title = re.sub(r"^\d+[-_\s]+", "", input_file.stem).strip() or input_file.stem
        sections: list[str] = [f"# {title}", ""]
        rendered = 0
        for idx, page in enumerate(doc, 1):
            fragment = _pdf_render_page(page, body_size, idx)
            if fragment:
                sections += [fragment, ""]
                rendered += 1
        body = "\n".join(sections).rstrip() + "\n"
        body = re.sub(r"\n{3,}", "\n\n", body)
        out_file.parent.mkdir(parents=True, exist_ok=True)
        out_file.write_text(body, encoding="utf-8")
        return {"pages": len(doc), "rendered_pages": rendered}
    finally:
        doc.close()


# Dispatcher -------------------------------------------------------------

def detect_format(input_file: Path) -> str:
    suffix = input_file.suffix.lower()
    if suffix in CSV_FORMATS:
        return "csv"
    if suffix in TSV_FORMATS:
        return "tsv"
    if suffix in TEXT_FORMATS:
        return "text"
    if suffix in MARKDOWN_FORMATS:
        return "markdown"
    if suffix in EXCEL_FORMATS:
        return "excel"
    if suffix in LEGACY_EXCEL_FORMATS:
        return "excel-legacy"
    if suffix in PDF_FORMATS:
        return "pdf"
    return "unknown"


def convert_to_markdown(
    input_path: str | Path,
    output_path: str | Path | None = None,
    project_path: str | Path | None = None,
    max_rows: int = 0,
    max_cols: int = 0,
) -> dict[str, Any]:
    input_file = Path(input_path)
    if not input_file.exists():
        raise FileNotFoundError(f"input file not found: {input_path}")
    if not input_file.is_file():
        raise FileNotFoundError(f"not a regular file: {input_path}")
    fmt = detect_format(input_file)
    if fmt == "unknown":
        raise ValueError(
            f"unsupported format '{input_file.suffix}'. "
            f"Supported: {', '.join(sorted(ALL_SUPPORTED))}"
        )
    if fmt == "excel-legacy":
        raise ValueError(
            "legacy .xls format is not supported. Resave as .xlsx and retry."
        )
    if max_rows < 0 or max_cols < 0:
        raise ValueError("--max-rows and --max-cols must be zero or positive")

    out_file = _resolve_output_path(
        input_file,
        str(output_path) if output_path is not None else None,
        str(project_path) if project_path is not None else None,
    )
    print(f"[INFO] Converting {fmt} source: {input_file.name}")
    if fmt in ("csv", "tsv"):
        extra = _convert_csv(input_file, out_file, max_rows, max_cols, fmt == "tsv")
    elif fmt == "text":
        extra = _convert_text(input_file, out_file)
    elif fmt == "markdown":
        extra = _convert_markdown(input_file, out_file)
    elif fmt == "excel":
        extra = _convert_excel(input_file, out_file, max_rows, max_cols)
    elif fmt == "pdf":
        extra = _convert_pdf(input_file, out_file)
    else:  # pragma: no cover
        raise ValueError(f"unhandled format: {fmt}")
    print(f"[OK] Wrote {out_file} ({_format_size(out_file.stat().st_size)})")
    return _emit_summary(out_file, input_file, fmt, extra)


def _iter_batch_inputs(directory: Path, recursive: bool) -> Iterable[Path]:
    pattern = "**/*" if recursive else "*"
    for path in sorted(directory.glob(pattern)):
        if path.is_file() and path.suffix.lower() in ALL_SUPPORTED:
            yield path


def batch_convert(
    input_dir: str | Path,
    project_path: str | Path | None,
    recursive: bool = True,
    max_rows: int = 0,
    max_cols: int = 0,
) -> list[dict[str, Any]]:
    directory = Path(input_dir)
    if not directory.exists() or not directory.is_dir():
        raise FileNotFoundError(f"directory not found: {input_dir}")
    summaries: list[dict[str, Any]] = []
    for input_file in _iter_batch_inputs(directory, recursive):
        try:
            summaries.append(convert_to_markdown(
                input_file, None, project_path, max_rows, max_cols,
            ))
        except (FileNotFoundError, ValueError, RuntimeError, KeyError) as exc:
            _print_error(f"{input_file}: {exc}")
            summaries.append({
                "status": "failed",
                "input": str(input_file),
                "error": str(exc),
            })
    return summaries


# CLI --------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="source_to_md",
        description=(
            "Unified source converter for visio-master inputs "
            "(CSV / TSV / TXT / MD / XLSX / PDF -> Markdown)."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python source_to_md.py convert org_roster.xlsx --project ./drawing
  python source_to_md.py convert spec.pdf -o sources/spec.md
  python source_to_md.py batch ./inputs --project ./drawing
  python source_to_md.py inspect data.csv

Output convention:
  <project_path>/sources/<input.stem>.md  (when --project is given)
  <input>.md                              (otherwise)
""",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    p = sub.add_parser("convert", help="Convert a single source file.")
    p.add_argument("input")
    p.add_argument("-o", "--output", help="Explicit output Markdown path")
    p.add_argument("--project", help="Project root; output to <project>/sources/")
    p.add_argument("--max-rows", type=int, default=0)
    p.add_argument("--max-cols", type=int, default=0)

    p = sub.add_parser("batch", help="Convert every supported file in a directory.")
    p.add_argument("input_dir")
    p.add_argument("--project")
    p.add_argument("--no-recursive", action="store_true")
    p.add_argument("--max-rows", type=int, default=0)
    p.add_argument("--max-cols", type=int, default=0)

    p = sub.add_parser("inspect", help="Probe a file's detected format.")
    p.add_argument("input")
    return parser


def _cmd_convert(args: argparse.Namespace) -> int:
    try:
        convert_to_markdown(
            args.input, args.output, args.project, args.max_rows, args.max_cols,
        )
    except (FileNotFoundError, ValueError, RuntimeError) as exc:
        _print_error(str(exc))
        return 1
    except KeyError as exc:
        _print_error(f"missing key: {exc}")
        return 2
    except Exception as exc:  # noqa: BLE001
        if _HAS_PYWIN32 and isinstance(exc, pywintypes.com_error):  # type: ignore[union-attr]
            _print_error(f"COM error during conversion: {exc}")
            return 2
        _print_error(f"unexpected error: {exc}")
        return 2
    return 0


def _cmd_batch(args: argparse.Namespace) -> int:
    try:
        summaries = batch_convert(
            args.input_dir, args.project, not args.no_recursive,
            args.max_rows, args.max_cols,
        )
    except FileNotFoundError as exc:
        _print_error(str(exc))
        return 1
    failures = [s for s in summaries if s.get("status") == "failed"]
    print(json.dumps({
        "status": "ok" if not failures else "partial",
        "converted": len(summaries) - len(failures),
        "failed": len(failures),
        "results": summaries,
    }, ensure_ascii=False))
    return 0 if not failures else 2


def _cmd_inspect(args: argparse.Namespace) -> int:
    input_file = Path(args.input)
    if not input_file.exists():
        _print_error(f"file not found: {args.input}")
        return 1
    fmt = detect_format(input_file)
    deps_ok = True
    notes: list[str] = []
    if fmt == "excel" and not _HAS_OPENPYXL:
        deps_ok = False
        notes.append("openpyxl missing — install with: pip install 'openpyxl>=3.1'")
    if fmt == "pdf" and not _HAS_FITZ:
        deps_ok = False
        notes.append("PyMuPDF missing — install with: pip install 'PyMuPDF>=1.24'")
    if fmt == "excel-legacy":
        notes.append("resave .xls as .xlsx before converting")
    if fmt == "unknown":
        notes.append(f"unsupported extension: {input_file.suffix}")
    status = "ok" if deps_ok and fmt not in ("unknown", "excel-legacy") else "blocked"
    print(json.dumps({
        "status": status,
        "input": str(input_file),
        "format": fmt,
        "size_bytes": input_file.stat().st_size,
        "deps_ok": deps_ok,
        "notes": notes,
    }, ensure_ascii=False))
    return 0


def main(argv: list[str] | None = None) -> int:
    args = _build_parser().parse_args(argv)
    if args.command == "convert":
        return _cmd_convert(args)
    if args.command == "batch":
        return _cmd_batch(args)
    if args.command == "inspect":
        return _cmd_inspect(args)
    return 1  # pragma: no cover - argparse exits earlier


if __name__ == "__main__":
    sys.exit(main())
